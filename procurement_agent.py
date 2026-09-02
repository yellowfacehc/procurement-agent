# -*- coding: utf-8 -*-
"""
采购意向快速获取工具
- 数据源：重庆市政府采购网
- 地区：重庆市九龙坡区、大渡口区
- 功能：日期范围筛选、关键词搜索、Excel导出
"""

import io
import time
import requests
from datetime import datetime, timedelta
from flask import Flask, request, jsonify, send_file, render_template_string
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Playwright为可选依赖，用于绕过中国政府采购网反爬
try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False
    print("[警告] 未安装playwright，中国政府采购网数据将不可用。")
    print("       安装命令: pip install playwright && playwright install chromium")

app = Flask(__name__)

# ============ 常量配置 ============
REGIONS = {
    "九龙坡区": "130117562645086212",
    "大渡口区": "130117562645086213",
    "重庆市本级": "130117562645086249",  # 监狱、大学等市级单位注册在此
    "高新区": "758829616340365312",  # 高新区政务服务中心等单位注册在此
}

# 中国政府采购网 - 重庆市zoneId
CQ_ZONE_ID = "500000"

# 关键词过滤（用于中国政府采购网结果中筛选九龙坡和大渡口）
TARGET_DISTRICTS = ["九龙坡", "大渡口"]

# 位于九龙坡区但名称不含"九龙坡"的单位（根据实际数据整理，可补充）
JIULONGPO_EXTRA_UNITS = [
    "重庆医科大学附属康复医院",
    "重庆市大坪监狱",
    "重庆市红光中学",
    "重庆市育才中学校",
    "重庆开放大学",
    "重庆工商职业学院",
    "重庆电力高等专科学校",
    "重庆高新技术产业开发区政务服务和社会事务中心",
    "重庆高新区政务服务和社会事务中心",
]

# 位于大渡口区但名称不含"大渡口"的单位
DADUKOU_EXTRA_UNITS = [
    "重庆市旅游学校",
]

def match_district(title, org_name):
    """判断标题或单位名称是否属于九龙坡区或大渡口区"""
    # 先检查区名关键词
    for d in TARGET_DISTRICTS:
        if d in title or d in org_name:
            return d + "区" if not d.endswith("区") else d
    # 再检查额外单位列表
    for u in JIULONGPO_EXTRA_UNITS:
        if u in org_name or u in title:
            return "九龙坡区"
    for u in DADUKOU_EXTRA_UNITS:
        if u in org_name or u in title:
            return "大渡口区"
    return None

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
}


# ============ 爬虫函数 ============
def search_chongqing(start_date, end_date, keyword=""):
    """爬取重庆市政府采购网采购意向（使用新API，Render优化版）"""
    import time as _time
    _start_time = _time.time()
    _MAX_TIME = 25  # 总体最大爬取时间（秒），Render超时30秒留余量

    results = []
    url = "https://www.ccgp-chongqing.gov.cn/yw-gateway/demand/demand/front"
    headers = {
        **HEADERS,
        "Referer": "https://www.ccgp-chongqing.gov.cn/stock-resources/front/intentionList",
    }

    for district_name, region_id in REGIONS.items():
        # 总体时间检查
        if _time.time() - _start_time > _MAX_TIME:
            print(f"[重庆网] 达到总体时间限制{_MAX_TIME}秒，停止后续region")
            break

        page = 1
        page_size = 50
        all_items = []

        while True:
            # 总体时间检查
            if _time.time() - _start_time > _MAX_TIME:
                print(f"[重庆网-{district_name}] 达到总体时间限制，停止分页")
                break

            params = {
                "type": 2,
                "page": page,
                "pageSize": page_size,
                "createRegionId": region_id,
                "__platDomain__": "www.ccgp-chongqing.gov.cn",
            }

            try:
                r = requests.get(url, params=params, headers=headers, timeout=10)
                r.raise_for_status()
                data = r.json()
                items = data.get("data", [])
                if not items:
                    break
                all_items.extend(items)
                total = int(data.get("count", 0))

                # 优化：如果当前页最早的数据已经早于开始日期，提前终止
                if start_date and items:
                    earliest_ts = min(item.get("createTime", 0) for item in items)
                    if earliest_ts:
                        earliest_date = datetime.fromtimestamp(earliest_ts / 1000).strftime("%Y-%m-%d")
                        if earliest_date < start_date:
                            print(f"[重庆网-{district_name}] 已爬取到{earliest_date}，早于开始日期{start_date}，提前终止")
                            break

                if page * page_size >= total:
                    break
                page += 1
                # Render环境下去掉sleep，加快速度
            except Exception as e:
                print(f"[重庆网-{district_name}] 第{page}页爬取失败: {e}")
                break

        # 在Python端做日期、关键词和地区筛选
        for item in all_items:
            title = item.get("title", "")
            org_name = item.get("createOrgName", item.get("budgetOrgName", ""))
            create_ts = item.get("createTime", 0)
            publish_date = datetime.fromtimestamp(create_ts / 1000).strftime("%Y-%m-%d") if create_ts else ""

            # 日期过滤
            if start_date and publish_date < start_date:
                continue
            if end_date and publish_date > end_date:
                continue

            # 关键词过滤（匹配标题或单位名称）
            if keyword and keyword not in title and keyword not in org_name:
                continue

            # 地区判断：九龙坡/大渡口区级数据直接保留，其他（市级、高新区）需要用match_district过滤
            if district_name in ("重庆市本级", "高新区"):
                matched = match_district(title, org_name)
                if not matched:
                    continue
                item_district = matched
            else:
                item_district = district_name

            results.append({
                "source": "重庆市政府采购网",
                "district": item_district,
                "title": title,
                "org_name": org_name,
                "publish_date": publish_date,
                "detail_url": f"https://www.ccgp-chongqing.gov.cn/stock-resources/front/intentionView?id={item.get('id', '')}",
                "biz_id": item.get("id", ""),
                "budget": item.get("money", ""),
                "depict": item.get("depict", ""),
            })

    return results


def search_ccgp(start_date, end_date, keyword=""):
    """使用Playwright爬取中国政府采购网（绕过反爬机制）"""
    if not PLAYWRIGHT_AVAILABLE:
        print("[中国政府采购网] 跳过：未安装playwright")
        return []

    results = []

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                locale="zh-CN",
            )
            page = context.new_page()

            # 访问页面建立会话
            page.goto("http://cgyx.ccgp.gov.cn/cgyx/pub/pubSearch", wait_until="networkidle", timeout=30000)
            time.sleep(2)

            page_no = 1
            page_size = 10
            total = None

            while True:
                # 通过页面内fetch调用API（使用页面的Cookie和上下文）
                js_code = f"""
                async () => {{
                    const formData = new URLSearchParams();
                    formData.append('releaseStar', '{start_date or ""}');
                    formData.append('releaseEnd', '{end_date or ""}');
                    formData.append('title', '{keyword or ""}');
                    formData.append('releaseUnitName', '');
                    formData.append('zoneId', '{CQ_ZONE_ID}');
                    formData.append('type', '1');
                    formData.append('pageSize', '{page_size}');
                    formData.append('pageNo', '{page_no}');

                    const resp = await fetch('/cgyx/pub/pubSearchData', {{
                        method: 'POST',
                        body: formData,
                        headers: {{
                            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
                            'X-Requested-With': 'XMLHttpRequest'
                        }}
                    }});
                    return await resp.json();
                }}
                """
                try:
                    resp = page.evaluate(js_code)
                except Exception as e:
                    print(f"[中国政府采购网] 第{page_no}页调用失败: {e}")
                    break

                if total is None:
                    total = int(resp.get("total", 0))
                    print(f"[中国政府采购网] 总计{total}条，开始分页爬取...")

                rows = resp.get("rows", [])
                if not rows:
                    break

                for row in rows:
                    title = row.get("groupName", "")
                    org_name = row.get("releaseUnitName", "")
                    # 过滤九龙坡和大渡口（含名称不含区名的所属单位）
                    matched_district = match_district(title, org_name)
                    if not matched_district:
                        continue

                    publish_date = row.get("releaseDate", "")[:10]
                    # 日期过滤
                    if start_date and publish_date < start_date:
                        continue
                    if end_date and publish_date > end_date:
                        continue
                    # 关键词过滤
                    if keyword and keyword not in title and keyword not in org_name:
                        continue

                    results.append({
                        "source": "中国政府采购网",
                        "district": matched_district,
                        "title": title,
                        "org_name": org_name,
                        "publish_date": publish_date,
                        "detail_url": f"http://cgyx.ccgp.gov.cn/cgyx/pub/details?groupId={row.get('groupId', '')}",
                        "biz_id": row.get("groupId", ""),
                        "budget": "",
                        "depict": "",
                    })

                if page_no * page_size >= total:
                    break
                page_no += 1
                time.sleep(0.5)

            browser.close()
    except Exception as e:
        print(f"[中国政府采购网] Playwright爬取异常: {e}")

    print(f"[中国政府采购网] 爬取完成，共{len(results)}条九龙坡/大渡口数据")
    return results


def do_search(start_date, end_date, keyword=""):
    """执行搜索，只爬取重庆市政府采购网"""
    stats = {"cq_raw": 0}

    # 只爬取重庆市政府采购网
    try:
        cq_results = search_chongqing(start_date, end_date, keyword)
        stats["cq_raw"] = len(cq_results)
    except Exception as e:
        print(f"[重庆网] 爬取异常: {e}")
        cq_results = []
        stats["cq_error"] = str(e)

    # 按发布日期降序排序
    cq_results.sort(key=lambda x: x.get("publish_date", ""), reverse=True)

    return cq_results, stats


# ============ 路由 ============
@app.route("/")
def index():
    return render_template_string(HTML_TEMPLATE, playwright_available=PLAYWRIGHT_AVAILABLE)


@app.route("/api/search", methods=["POST"])
def api_search():
    data = request.get_json() or {}
    start_date = data.get("start_date", "").strip()
    end_date = data.get("end_date", "").strip()
    keyword = data.get("keyword", "").strip()
    districts = data.get("districts", [])

    # 默认日期：最近7天
    if not start_date:
        start_date = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")

    try:
        results, stats = do_search(start_date, end_date, keyword)
        # 地区筛选
        if districts:
            results = [r for r in results if any(d in r["district"] for d in districts)]
        return jsonify({
            "success": True,
            "count": len(results),
            "start_date": start_date,
            "end_date": end_date,
            "keyword": keyword,
            "districts": districts,
            "results": results,
            "stats": stats,
        })
    except Exception as e:
        import traceback
        print(f"[API错误] {traceback.format_exc()}")
        return jsonify({"success": False, "error": f"搜索出错：{str(e)}"}), 200


@app.route("/api/export", methods=["POST"])
def api_export():
    data = request.get_json() or {}
    start_date = data.get("start_date", "").strip()
    end_date = data.get("end_date", "").strip()
    keyword = data.get("keyword", "").strip()
    districts = data.get("districts", [])

    if not start_date:
        start_date = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")

    results, stats = do_search(start_date, end_date, keyword)
    # 地区筛选
    if districts:
        results = [r for r in results if any(d in r["district"] for d in districts)]

    # 生成Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "采购意向"

    # 表头样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["序号", "地区", "采购项目名称", "发布单位", "发布日期", "预算金额(万元)", "详情链接"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 数据行
    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)
    for i, item in enumerate(results, 1):
        row = i + 1
        values = [
            i,
            item["district"],
            item["title"],
            item["org_name"],
            item["publish_date"],
            item.get("budget", ""),
            item["detail_url"],
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    # 列宽
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 60

    # 冻结首行
    ws.freeze_panes = "A2"

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"采购意向_{start_date}_{end_date}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============ 前端模板 ============
HTML_TEMPLATE = r"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>采购意向快速获取工具</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: "Microsoft YaHei", "微软雅黑", -apple-system, sans-serif;
            background: #f0f2f5;
            color: #333;
            min-height: 100vh;
        }
        .container { max-width: 1200px; margin: 0 auto; padding: 20px; }
        .header {
            background: linear-gradient(135deg, #1a73e8, #0d47a1);
            color: white;
            padding: 24px 30px;
            border-radius: 12px;
            margin-bottom: 20px;
            box-shadow: 0 4px 12px rgba(26,115,232,0.3);
        }
        .header h1 { font-size: 24px; margin-bottom: 6px; }
        .header p { font-size: 14px; opacity: 0.9; }
        .status-tip {
            background: #fff3cd;
            border-left: 4px solid #ffc107;
            padding: 10px 16px;
            border-radius: 6px;
            margin-top: 12px;
            font-size: 13px;
            color: #856404;
        }
        .search-panel {
            background: white;
            padding: 24px;
            border-radius: 12px;
            margin-bottom: 20px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        }
        .form-row {
            display: flex;
            gap: 16px;
            align-items: flex-end;
            flex-wrap: wrap;
        }
        .form-group { display: flex; flex-direction: column; gap: 6px; }
        .form-group label {
            font-size: 13px;
            color: #666;
            font-weight: 500;
        }
        .form-group input {
            padding: 10px 14px;
            border: 1px solid #d9d9d9;
            border-radius: 8px;
            font-size: 14px;
            outline: none;
            transition: border-color 0.2s;
            width: 180px;
        }
        .form-group input:focus { border-color: #1a73e8; box-shadow: 0 0 0 3px rgba(26,115,232,0.1); }
        .form-group.keyword input { width: 280px; }
        .btn {
            padding: 10px 24px;
            border: none;
            border-radius: 8px;
            font-size: 14px;
            cursor: pointer;
            transition: all 0.2s;
            font-weight: 500;
        }
        .btn-primary {
            background: #1a73e8;
            color: white;
        }
        .btn-primary:hover { background: #1557b0; }
        .btn-primary:disabled { background: #9cc3f5; cursor: not-allowed; }
        .btn-success {
            background: #52c41a;
            color: white;
        }
        .btn-success:hover { background: #389e0d; }
        .btn-success:disabled { background: #b7e48a; cursor: not-allowed; }
        .region-tags {
            margin-top: 12px;
            display: flex;
            gap: 8px;
            align-items: center;
        }
        .region-tags span { font-size: 13px; color: #666; }
        .checkbox-label {
            display: inline-flex;
            align-items: center;
            gap: 4px;
            font-size: 13px;
            color: #333;
            cursor: pointer;
            padding: 4px 10px;
            background: #f5f5f5;
            border-radius: 6px;
            transition: background 0.2s;
        }
        .checkbox-label:hover { background: #e8e8e8; }
        .checkbox-label input { width: auto; margin: 0; cursor: pointer; }
        .tag {
            display: inline-block;
            padding: 4px 12px;
            background: #e6f4ff;
            color: #1a73e8;
            border-radius: 12px;
            font-size: 12px;
            font-weight: 500;
        }
        .stats-bar {
            display: flex;
            gap: 20px;
            margin-bottom: 16px;
            flex-wrap: wrap;
        }
        .stat-card {
            background: white;
            padding: 16px 24px;
            border-radius: 10px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
            flex: 1;
            min-width: 150px;
        }
        .stat-card .label { font-size: 13px; color: #888; margin-bottom: 4px; }
        .stat-card .value { font-size: 24px; font-weight: 700; color: #1a73e8; }
        .stat-card .value.green { color: #52c41a; }
        .stat-card .value.orange { color: #fa8c16; }
        .results-panel {
            background: white;
            border-radius: 12px;
            overflow: hidden;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        }
        .results-header {
            padding: 16px 24px;
            border-bottom: 1px solid #f0f0f0;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .results-header h3 { font-size: 16px; }
        table { width: 100%; border-collapse: collapse; }
        thead { background: #fafafa; }
        th {
            padding: 12px 16px;
            text-align: left;
            font-size: 13px;
            color: #666;
            font-weight: 600;
            border-bottom: 2px solid #f0f0f0;
            white-space: nowrap;
        }
        td {
            padding: 12px 16px;
            font-size: 13px;
            border-bottom: 1px solid #f5f5f5;
            vertical-align: middle;
        }
        td:nth-child(1), td:nth-child(2), td:nth-child(4), td:nth-child(5), td:nth-child(6) {
            white-space: nowrap;
        }
        tr:hover { background: #f9fbff; }
        .source-badge {
            display: inline-block;
            padding: 2px 8px;
            border-radius: 4px;
            font-size: 11px;
            font-weight: 500;
        }
        .source-ccgp { background: #fff7e6; color: #fa8c16; }
        .source-cq { background: #e6f4ff; color: #1a73e8; }
        .district-badge {
            display: inline-block;
            padding: 3px 8px;
            background: #f6ffed;
            color: #52c41a;
            border-radius: 4px;
            font-size: 12px;
            white-space: nowrap;
        }
        .title-link {
            color: #1a73e8;
            text-decoration: none;
            line-height: 1.5;
        }
        .title-link:hover { text-decoration: underline; }
        .empty-state {
            padding: 60px 20px;
            text-align: center;
            color: #999;
        }
        .empty-state .icon { font-size: 48px; margin-bottom: 12px; }
        .loading {
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 40px;
            color: #1a73e8;
            font-size: 15px;
        }
        .spinner {
            width: 20px;
            height: 20px;
            border: 3px solid #e6f4ff;
            border-top-color: #1a73e8;
            border-radius: 50%;
            animation: spin 0.8s linear infinite;
            margin-right: 10px;
        }
        @keyframes spin { to { transform: rotate(360deg); } }
        .error-msg {
            background: #fff2f0;
            border: 1px solid #ffccc7;
            color: #cf1322;
            padding: 12px 20px;
            border-radius: 8px;
            margin-bottom: 16px;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>采购意向快速获取工具</h1>
            <p>数据来源：重庆市政府采购网 | 地区：九龙坡区 · 大渡口区</p>
        </div>

        <div class="search-panel">
            <div class="form-row">
                <div class="form-group">
                    <label>开始日期</label>
                    <input type="date" id="startDate">
                </div>
                <div class="form-group">
                    <label>结束日期</label>
                    <input type="date" id="endDate">
                </div>
                <div class="form-group keyword">
                    <label>关键词（留空搜索全部）</label>
                    <input type="text" id="keyword" placeholder="如：医院、信息化、物业...">
                </div>
                <button class="btn btn-primary" id="searchBtn" onclick="doSearch()">搜索</button>
                <button class="btn btn-success" id="exportBtn" onclick="doExport()" disabled>导出Excel</button>
            </div>
            <div class="region-tags">
                <span>筛选地区：</span>
                <label class="checkbox-label"><input type="checkbox" id="districtJLP" value="九龙坡" checked> 九龙坡区</label>
                <label class="checkbox-label"><input type="checkbox" id="districtDDK" value="大渡口" checked> 大渡口区</label>
            </div>
        </div>

        <div id="errorBox"></div>
        <div id="statsBox"></div>

        <div class="results-panel">
            <div class="results-header">
                <h3>搜索结果</h3>
                <span id="resultCount" style="font-size:13px;color:#888;"></span>
            </div>
            <div id="resultsBody">
                <div class="empty-state">
                    <div class="icon">&#128269;</div>
                    <p>请选择日期范围并点击搜索</p>
                </div>
            </div>
        </div>
    </div>

    <script>
        let lastResults = [];

        // 初始化默认日期（最近7天）
        function initDates() {
            const today = new Date();
            const weekAgo = new Date(today.getTime() - 7 * 24 * 60 * 60 * 1000);
            document.getElementById('endDate').value = today.toISOString().split('T')[0];
            document.getElementById('startDate').value = weekAgo.toISOString().split('T')[0];
        }
        initDates();

        function showError(msg) {
            document.getElementById('errorBox').innerHTML =
                '<div class="error-msg">' + msg + '</div>';
        }
        function clearError() {
            document.getElementById('errorBox').innerHTML = '';
        }

        async function doSearch() {
            clearError();
            const startDate = document.getElementById('startDate').value;
            const endDate = document.getElementById('endDate').value;
            const keyword = document.getElementById('keyword').value.trim();

            // 获取选中的地区
            const districts = [];
            if (document.getElementById('districtJLP').checked) districts.push('九龙坡');
            if (document.getElementById('districtDDK').checked) districts.push('大渡口');

            if (!startDate || !endDate) {
                showError('请选择开始日期和结束日期');
                return;
            }
            if (startDate > endDate) {
                showError('开始日期不能晚于结束日期');
                return;
            }
            if (districts.length === 0) {
                showError('请至少选择一个地区');
                return;
            }

            const btn = document.getElementById('searchBtn');
            const exportBtn = document.getElementById('exportBtn');
            btn.disabled = true;
            btn.textContent = '搜索中...';
            exportBtn.disabled = true;

            document.getElementById('resultsBody').innerHTML =
                '<div class="loading"><div class="spinner"></div>正在爬取网站数据，请稍候...</div>';
            document.getElementById('statsBox').innerHTML = '';

            try {
                const resp = await fetch('/api/search', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({start_date: startDate, end_date: endDate, keyword: keyword, districts: districts})
                });
                const data = await resp.json();

                if (!data.success) {
                    throw new Error(data.error || '搜索失败');
                }

                lastResults = data.results;
                renderResults(data);
                exportBtn.disabled = data.count === 0;
            } catch (e) {
                showError('搜索失败：' + e.message);
                document.getElementById('resultsBody').innerHTML =
                    '<div class="empty-state"><div class="icon">&#9888;</div><p>搜索失败，请重试</p></div>';
            } finally {
                btn.disabled = false;
                btn.textContent = '搜索';
            }
        }

        function renderResults(data) {
            const results = data.results;
            document.getElementById('resultCount').textContent = '共 ' + data.count + ' 条';

            if (results.length === 0) {
                document.getElementById('resultsBody').innerHTML =
                    '<div class="empty-state"><div class="icon">&#128269;</div><p>未找到符合条件的采购意向</p></div>';
                document.getElementById('statsBox').innerHTML = '';
                return;
            }

            // 统计
            const cqCount = results.filter(r => r.source === '重庆市政府采购网').length;
            const jlpCount = results.filter(r => r.district.includes('九龙坡')).length;
            const ddkCount = results.filter(r => r.district.includes('大渡口')).length;

            // 原始数量和错误信息
            const stats = data.stats || {};
            const cqRaw = stats.cq_raw || 0;

            let cqDisplay = cqCount;
            if (cqRaw > cqCount) {
                cqDisplay = cqCount + ' <span style="font-size:11px;color:#999;">(原始' + cqRaw + '条)</span>';
            }

            document.getElementById('statsBox').innerHTML =
                '<div class="stats-bar">' +
                '<div class="stat-card"><div class="label">总计</div><div class="value">' + data.count + '</div></div>' +
                '<div class="stat-card"><div class="label">重庆市政府采购网</div><div class="value">' + cqDisplay + '</div></div>' +
                '<div class="stat-card"><div class="label">九龙坡区</div><div class="value green">' + jlpCount + '</div></div>' +
                '<div class="stat-card"><div class="label">大渡口区</div><div class="value green">' + ddkCount + '</div></div>' +
                '</div>';

            // 表格
            let html = '<table><thead><tr>' +
                '<th style="width:50px;">序号</th>' +
                '<th style="width:80px;">地区</th>' +
                '<th>采购项目名称</th>' +
                '<th style="width:220px;">发布单位</th>' +
                '<th style="width:100px;">发布日期</th>' +
                '<th style="width:100px;">预算(万元)</th>' +
                '</tr></thead><tbody>';

            results.forEach((item, idx) => {
                html += '<tr>' +
                    '<td>' + (idx + 1) + '</td>' +
                    '<td><span class="district-badge">' + item.district + '</span></td>' +
                    '<td><a class="title-link" href="' + item.detail_url + '" target="_blank">' + escapeHtml(item.title) + '</a></td>' +
                    '<td>' + escapeHtml(item.org_name) + '</td>' +
                    '<td>' + item.publish_date + '</td>' +
                    '<td>' + (item.budget || '-') + '</td>' +
                    '</tr>';
            });

            html += '</tbody></table>';
            document.getElementById('resultsBody').innerHTML = html;
        }

        function escapeHtml(text) {
            const div = document.createElement('div');
            div.textContent = text;
            return div.innerHTML;
        }

        async function doExport() {
            if (lastResults.length === 0) return;
            const startDate = document.getElementById('startDate').value;
            const endDate = document.getElementById('endDate').value;
            const keyword = document.getElementById('keyword').value.trim();
            const districts = [];
            if (document.getElementById('districtJLP').checked) districts.push('九龙坡');
            if (document.getElementById('districtDDK').checked) districts.push('大渡口');

            const btn = document.getElementById('exportBtn');
            btn.disabled = true;
            btn.textContent = '导出中...';

            try {
                const resp = await fetch('/api/export', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({start_date: startDate, end_date: endDate, keyword: keyword, districts: districts})
                });
                if (!resp.ok) throw new Error('导出失败');
                const blob = await resp.blob();
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = '采购意向_' + startDate + '_' + endDate + '.xlsx';
                document.body.appendChild(a);
                a.click();
                document.body.removeChild(a);
                window.URL.revokeObjectURL(url);
            } catch (e) {
                showError('导出失败：' + e.message);
            } finally {
                btn.disabled = false;
                btn.textContent = '导出Excel';
            }
        }

        // 回车搜索
        document.getElementById('keyword').addEventListener('keypress', function(e) {
            if (e.key === 'Enter') doSearch();
        });
    </script>
</body>
</html>
"""


if __name__ == "__main__":
    import socket
    # 获取本机局域网IP
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
    except:
        local_ip = "127.0.0.1"

    print("=" * 55)
    print("  采购意向快速获取工具")
    print("  数据源：重庆市政府采购网")
    print("  地区：九龙坡区 · 大渡口区")
    print("=" * 55)
    print(f"  本机访问: http://127.0.0.1:5000")
    print(f"  局域网访问: http://{local_ip}:5000")
    print("  (同一网络下的同事可通过局域网地址访问)")
    print("=" * 55)
    app.run(host="0.0.0.0", port=5000, debug=False)
